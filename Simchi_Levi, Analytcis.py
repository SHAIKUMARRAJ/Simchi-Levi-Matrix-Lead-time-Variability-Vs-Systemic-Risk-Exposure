import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
import matplotlib.pyplot as plt
import seaborn as sns

def run_simchi_levi_pipeline():
    print("🚀 Initializing David Simchi-Levi Procurement Risk Pipeline...")
    
    # 1. Load Data from Source Database
    file_path = "IT_Procurement_Project.xlsx"
    try:
        df_po = pd.read_excel(file_path, sheet_name="Purchase_Orders")
        df_sup = pd.read_excel(file_path, sheet_name="Suppliers")
    except Exception as e:
        print(f"❌ Error loading 'IT_Procurement_Project.xlsx'. Ensure it is in the same directory. Details: {e}")
        return

    df_po['PO_Date'] = pd.to_datetime(df_po['PO_Date'])

    # 2. Compute Integrated Supply Chain Risk Metrics
    print("📈 Calculating Lead-Time Variability (Sigma) and Daily Spend Velocity...")
    
    # Calculate Lead-Time standard deviation (Delivery Unpredictability)
    lt_stats = df_po.groupby('Supplier_ID')['Lead_Time'].agg(['mean', 'std']).reset_index()
    lt_stats.columns = ['Supplier_ID', 'Average_Lead_Time', 'Lead_Time_Variability_Sigma']
    lt_stats['Lead_Time_Variability_Sigma'] = lt_stats['Lead_Time_Variability_Sigma'].fillna(0)

    # Calculate financial pipeline velocity per supplier
    total_days = (df_po['PO_Date'].max() - df_po['PO_Date'].min()).days
    supplier_spend = df_po.groupby('Supplier_ID')['Total_Amount'].sum().reset_index()
    supplier_spend['Daily_Spend_Velocity'] = supplier_spend['Total_Amount'] / total_days

    # Merge data sheets into the final analytical framework
    df_analysis = df_sup.merge(lt_stats, on='Supplier_ID', how='inner')
    df_analysis = df_analysis.merge(supplier_spend[['Supplier_ID', 'Daily_Spend_Velocity']], on='Supplier_ID', how='inner')
    
    # Calculate Time-to-Recovery (TTR) and Risk Exposure Index (REI)
    df_analysis['Simulated_TTR_Days'] = np.round(df_analysis['Risk_Score'] * 12) 
    df_analysis['Risk_Exposure_Index_REI'] = df_analysis['Daily_Spend_Velocity'] * df_analysis['Simulated_TTR_Days']
    df_analysis = df_analysis.sort_values(by='Risk_Exposure_Index_REI', ascending=False).reset_index(drop=True)

    # 3. Save Matplotlib Plot as a Static Image Asset for Documentation/LinkedIn
    print("📊 Generating High-Resolution Scatter Plot Image Asset...")
    plt.figure(figsize=(10, 6))
    sns.scatterplot(
        data=df_analysis, 
        x='Lead_Time_Variability_Sigma', 
        y='Risk_Exposure_Index_REI', 
        hue='Category',
        size='Average_Lead_Time',
        sizes=(40, 300),
        palette='deep',
        alpha=0.8
    )
    plt.title("Simchi-Levi Risk Matrix: Lead-Time Variability vs. Financial Risk Exposure (REI)", fontsize=11, fontweight='bold', pad=12)
    plt.xlabel("Lead-Time Variability (Sigma - Delivery Unpredictability)")
    plt.ylabel("Risk Exposure Index (REI - Financial Exposure During Outage)")
    plt.grid(True, linestyle='--', alpha=0.5)
    plt.tight_layout()
    
    image_output = "simchi_levi_matrix_chart.png"
    plt.savefig(image_output, dpi=300)
    plt.close()
    print(f"   -> Saved: '{image_output}'")

    # 4. Generate the Final Formatted Excel Workbook with Embedded Native Chart
    print("📑 Creating Formatted Executive Spreadsheet Report...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk_Register"
    ws.views.sheetView[0].showGridLines = True  # Keep default gridlines visible

    # Define Data Schema
    headers = ['Supplier_ID', 'Supplier_Name', 'Category', 'Country', 'Risk_Score', 
               'Average_Lead_Time', 'Lead_Time_Variability_Sigma', 'Daily_Spend_Velocity', 
               'Simulated_TTR_Days', 'Risk_Exposure_Index_REI']
    ws.append(headers)

    # Styling Elements
    navy_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    red_alert = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    # Format and paint Header Row
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = navy_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Populate Data Records
    for _, row in df_analysis.iterrows():
        ws.append([
            row['Supplier_ID'], row['Supplier_Name'], row['Category'], row['Country'], row['Risk_Score'],
            row['Average_Lead_Time'], row['Lead_Time_Variability_Sigma'], row['Daily_Spend_Velocity'],
            row['Simulated_TTR_Days'], row['Risk_Exposure_Index_REI']
        ])

    # Apply conditional styling, fonts, borders, and number formatting rules
    for row_idx in range(2, len(df_analysis) + 2):
        row_cells = ws[row_idx]
        
        # Center format structural keys
        row_cells[0].alignment = Alignment(horizontal="center")
        row_cells[4].alignment = Alignment(horizontal="center")
        row_cells[8].alignment = Alignment(horizontal="center")
        
        # Numeric column formatting
        row_cells[5].number_format = "0.0"     # Avg Lead Time
        row_cells[6].number_format = "0.00"    # Lead Time Variability (Sigma)
        row_cells[7].number_format = "$#,##0"  # Spend Velocity
        row_cells[9].number_format = "$#,##0"  # Risk Exposure Index (REI)
        
        # Alternating Zebra lines for readability
        if row_idx % 2 == 0:
            for cell in row_cells:
                cell.fill = zebra_fill
                
        # soft red highlighting for extreme risk nodes (REI > $15 Million)
        if row_cells[9].value > 15000000:
            row_cells[9].fill = red_alert

        for cell in row_cells:
            cell.font = font_regular
            cell.border = thin_border

    # 5. Build and Inject Interactive Native Excel Scatter Chart
    print("📊 Embedding Interactive Native Chart into Excel Sheet Sheet Grid...")
    chart = ScatterChart()
    chart.title = "Simchi-Levi Risk Matrix (Lead-Time vs. REI Exposure)"
    chart.style = 13
    chart.x_axis.title = 'Lead-Time Variability (Sigma)'
    chart.y_axis.title = 'Risk Exposure Index (REI)'

    # Set cell reference coordinates: X = Column 7 (Sigma), Y = Column 10 (REI)
    xvalues = Reference(ws, min_col=7, min_row=2, max_row=len(df_analysis) + 1)
    yvalues = Reference(ws, min_col=10, min_row=2, max_row=len(df_analysis) + 1)
    series = Series(yvalues, xvalues, title_from_data=False)
    
    # Configure markers (clean scatter points without a connective line plot line)
    series.marker.symbol = "circle"
    series.marker.size = 7
    series.graphicalProperties.line.noFill = True 

    chart.series.append(series)
    chart.width = 18
    chart.height = 11
    
    # Render chart at cell anchor location column L, row 2
    ws.add_chart(chart, "L2")

    # 6. Adjust dynamic column tracking boundaries
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    excel_output = "Simchi_Levi_Risk_Report_With_Chart.xlsx"
    wb.save(excel_output)
    print(f"   -> Saved: '{excel_output}'")
    print("✅ End-to-End Pipeline Completed Successfully!")

if __name__ == "__main__":
    run_simchi_levi_pipeline()